###########################################
## 
## Test Class for thermal laser
## SD Team 10: Robert Bara, Zari Grandy, Ezra Galapo, Tyiana Smith
## 
## Date:    4/4/2022
## Version: 1.0
##
## Description:
##  This python program allows for the thermal laser plantar test to run by prompting the user if
##  they would like to perform a test. The class creates an animal, performs a laser test, stores the withdrawal
##  times, as well as performs analysis such as standard dev, average, and variance calculations
##
## Usage:
##  Running: python3 test_class.py 
##
########################################

#imports
import animal_class as animal
from datetime import date   
from datetime import datetime
import os
from os import stat
from re import S
from sre_constants import MAXGROUPS             # Import date class from datetime module
from statistics import pstdev    # Import statistics to utilize mean, standard deviation, and variance calculations
from statistics import pvariance
from statistics import mean
from enum import Enum
#import RPi.GPIO as GPIO        # Import General-Purpose In/Out for RPI4 to control laser and photodiode
from time import time
from numpy import save           # Import time to calculate withdrawal time
import openpyxl                 # Import for exporting the animal's results to spreadsheets

#####################################################################
# Create a class to run a test and store all the animals information 
#####################################################################
class testAnimals:
    #No need for a constructor
    def __init__(self) -> None:
        pass
    ############################################################
    # Functions to Print results in various ways to the terminal
    ############################################################
    #Print the results of the Passed in Group number
    def printGroup(self, saveAnimals, group):
        #If the group's number matches the group argument, print Animal's information
        for i in saveAnimals:
            if(i.getGroup()==group): 
                i.printTime() 
        return
    #Prints the results of all groups in orderS
    def printAllGroups(self, saveAnimals):
        numOfGroups=self.maxGroup(saveAnimals)
        for i in range(numOfGroups+1):
            self.printGroup(saveAnimals,i)
        return
    #Function determines how many groups are there    
    def maxGroup(self,saveAnimals):
        numOfGroups=0
        for i in saveAnimals:
            if(i.group>numOfGroups):
                numOfGroups=i.group
        return numOfGroups

    #This function prints today's Test Animal's results
    def printTodaysResults(self, saveAnimals):
        print("\n--------------------------------------")
        print("Test Results:")
        print("--------------------------------------\n")
        #Print every Animal's information
        for i in saveAnimals:
            i.printTime() 
        return
    #####################################################
    # Functions for exporting results to spreadsheet
    #####################################################
    #This function creates a spreadsheet and labels the proper columns for the animal's data
    def exportSetup(self,saveAnimals):
        numOfGroups=self.maxGroup(saveAnimals)
        #Create a work book
        wb1=openpyxl.Workbook()
        #Create a worksheet within the workbook
        ws1=wb1.active
        ws1.title="Results"
        #initialize all the column names
        ws1.append(["Group","Num of Group Trials","Group Avg","Animal Name","Trial Num","Withdrawal Times","Avg Time",
                     "Stdev","Trial Variance","Test Date", "Start Test Time","End Test Time"])
        #prompt for an output file name
        workbookName=input("Save Output File As .xlsx?")
        wb1.save(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        #close the workbook
        wb1.close
        return workbookName #return the workbook name so we can write to it with the other export functions

    #This function exports the animals that were tested on the current day, to a spreadsheet
    def exportResults(self,saveAnimals,saveAnimalsLength):
        #Call excel setup function
        workbookName=self.exportSetup(saveAnimals)
        #Load the workbook that has been created
        wb1=openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        #Write Data to the current sheet
        ws1=wb1.active
        exportAnimals=saveAnimals
        #Loop through the saved animals
        for i in range(saveAnimalsLength):
            #remove the animal at the front of the list 
            currAnimal=exportAnimals.pop(0)
            #set the counter for the trial number
            trialNum=1
            #Loop through the Current Animal's withdrawal times
            for j in currAnimal.time:
                withdrawalTime=j
                group=currAnimal.getGroup()
                #Print the animal's attributes
                ws1.append([group,"n/a","n/a",currAnimal.getName(), trialNum, withdrawalTime, 
                            currAnimal.avg, currAnimal.standardDev, currAnimal.trialVariance
                            ,currAnimal.date, currAnimal.startTime,currAnimal.endTestTime])       
                trialNum+=1
        ws1.add_chart()
        # When exporting is done, save the workbook
        wb1.save(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        # Close the workbook and return
        wb1.close()
        return

    #Function for printing every group in order (Not finished yet)
    def exportAllGroups(self, saveAnimals,saveAnimalsLength,groupnum=0):
        #Call excel setup function
        workbookName=self.exportSetup(saveAnimals)
        #Load the workbook that has been created
        wb1=openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        #Write Data to the current sheet
        ws1=wb1.active
        exportAnimals=saveAnimals
        
        #Loop through the saved animals by group
        numOfGroups=self.maxGroup(exportAnimals)
        k=groupnum
        groupAvg=[]
        for k in range(numOfGroups+1):
            groupWithdrawlTimes=[]
            groupNumTrials=1
            for i in range(saveAnimalsLength):
                if(exportAnimals[i].getGroup()==k): 
                    #remove the animal at the front of the list 
                    currAnimal=exportAnimals[i]
                    #set the counter for the trial number
                    trialNum=1
                    analyzeTimes=[]
                    #Loop through the Current Animal's withdrawal times
                    for j in currAnimal.time:
                        analyzeTimes.append(j)
                        groupWithdrawlTimes.append(j)
                        #Print the animal's attributes
                        ws1.append([currAnimal.getGroup(),groupNumTrials,mean(groupWithdrawlTimes),currAnimal.getName(), trialNum, j, 
                                    mean(analyzeTimes), pstdev(analyzeTimes), pvariance(analyzeTimes)
                                    ,currAnimal.date, currAnimal.startTime, currAnimal.endTestTime])       
                        trialNum+=1
                        groupNumTrials+=1
        # When exporting is done, save the workbook
        wb1.save(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        # Close the workbook and return
        wb1.close()
        return

    def generateGraph():
        #check if when the last group entry for cat
        

        return
    
    ###############################################################################             
    #  Functions for starting the test and prompting for an animals information                          
    ###############################################################################    
    #Prompts the user for name and number to construct animal
    def promptAnimalInfo():
        #Create a animal with their information
        #Prompt for name of Animal
        name=input("Enter subject's Name: ")
        #Prompt for a valid integer group number
        while True:
            try:
                group = int(input("Enter subject's Group number: "))
                break
            except ValueError:
                print("Please enter Group number as an integer")  
            continue
        #Return Current Animal's name and number
        return name,group
    
    #This method is what starts the entire system to even run a test
    def startExperiment(self):
        saveAnimals=[]  #Create a list to store all animal's information
        saveAnimalsLength=0 #Keep track of the size of the list
       #Prompt user to start program
        print("\n*******Welcome to the Plantar Thermal Laser test.*******")
        beginProgram=input("Would you like to run a trial? (Yes/No): ")
        print(beginProgram)

        #Determine valid response, if so create a animal as an object
        while (beginProgram.lower()=="y")or (beginProgram.lower()=="yes"):
            #Create an animal data structure
            name,group=testAnimals.promptAnimalInfo()
            curr_animal=animal.animal(name,group)
            #Begin Testing a animal
            curr_animal.startTime=datetime.now().strftime("%H:%M:%S")  #get the timestamp when testing is beginning for cat
            trialNum=1 #Keeps track of trial of current animal
            runTrial=1 #Flag to determine which animal is being tested on
            #Run as many trials as wanted for this animal
            while (runTrial==1):
                #Perform a Trial, get withdrawal time
                print("Trial ",trialNum,":")
                curr_animal.trial()

                #Prompt to perform another trial
                nextTrial=input("Run another trial for this animal?(Yes/No) ")
                if (nextTrial.lower() == "y") or (nextTrial.lower()=="yes"): #Keep testing current animal
                    trialNum+=1
                    runTrial=1 
                if (nextTrial.lower() == "n") or (nextTrial.lower()=="no"): #Stop testing current animal
                    trialNum=0
                    runTrial=0
            #Print animal's information after testing 
            curr_animal.endTestTime=datetime.now().strftime("%H:%M:%S")  #get the timestamp when testing is over for cat
            curr_animal.printTime()        
            #Ask user to test another animal
            testAnotheranimal=input("\nTest a new animal?(Yes/No)")
            #If the User Answers Yes
            if (testAnotheranimal.lower() == "y") or (testAnotheranimal.lower() == "yes"): 
                saveAnimals.append(curr_animal) #add another animal to the list
                saveAnimalsLength+=1            #update the length of the list
                beginProgram="y" #Jump back to start of loop and create new animal
            #If the User Answers No
            if (testAnotheranimal.lower() == "n") or (testAnotheranimal.lower() == "no"): 
                saveAnimals.append(curr_animal) #add another animal to the list
                saveAnimalsLength+=1            #update the length of the list
                #self.printTodaysResults(saveAnimals)
                
                #Testing printing per group
                #self.printGroup(saveAnimals,0)
                #self.printGroup(saveAnimals,1)
                self.printAllGroups(saveAnimals)
                #self.exportResults(saveAnimals,saveAnimalsLength)
                self.exportAllGroups(saveAnimals, saveAnimalsLength)
                beginProgram="n" #Exit loop
                        
        #Exit program when testing is stopped
        if (beginProgram.lower() == "n") or (beginProgram.lower() == "no"):     
            print("*****Ending Testing*****")
        return

#########################################
#           Driver Function             #
#########################################
def main():
    #create a test and start the experiment
    test=testAnimals()
    test.startExperiment()
    #exit program
    return 0

##Main definition##
if __name__ == "__main__":
    main()    
