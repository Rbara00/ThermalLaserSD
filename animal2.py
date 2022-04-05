###########################################
## 
## Animal Class for thermal laser
## SD Team 10: Robert Bara, Zari Grandy, Ezra Galapo, Tyiana Smith
## 
## Date:    4/5/2022
## Version: 2.4
##
## Description:
##  This python file contains the class functions to create, obtain data, and analyze the
##  data for the thermal laser plantar test.
##
## Usage:
##  This program is invoked by test_class.py
##
########################################

############ Includes ##################
from datetime import date   
from datetime import datetime
import os
from os import stat
from re import S
import re
from sre_constants import MAXGROUPS             # Import date class from datetime module
from statistics import pstdev    # Import statistics to utilize mean, standard deviation, and variance calculations
from statistics import pvariance
from statistics import mean
from enum import Enum
#import RPi.GPIO as GPIO        # Import General-Purpose In/Out for RPI4 to control laser and photodiode
from time import time
from numpy import empty, save           # Import time to calculate withdrawal time
import openpyxl                 # Import for exporting the animal's results to spreadsheets
#############################################################
# Table Class
#############################################################
class resultTb:
    import animal2 as animal

    #constructor for a table of lists which hold a test animal as an element
    def __init__(self, size=100):
        self.size=size                              #dynamically allocated memory for the size of the table
        self.arr=[[] for i in range (self.size)]    #Creating the table as an array of lists
        self.filled=0                               #Keeps track of how man animals are "filled" in the table
        return
    #Takes an animal and inserts it into the results table based upon their group number
    def insert(self,curr_animal):
        group=curr_animal.getGroup()
        self.arr[group].append(curr_animal)
        self.filled+=1
        return
    #Prints all animals results within the passed in group number, in order of when they were stored into group
    def printIndex(self,group):
        #in the table at row with index corresponding to the group number, iterate across the list of animals
        for i in range(len(self.arr[group])):
            print(self.arr[group][i].printTime())   #Print the current counter's animal results
        return
    #print all results stored within the table in order of groups
    def printAll(self):
        for i in range(self.size):
            self.printIndex(i)
        return
############ animal Class #################
class animal:
    ##############################
    #Constructor for animal
    ##############################
    def __init__(self,name,group):
        self.name=name            #animal is named from user input
        self.time=[]              #create a list of all the animal's withdrawal times
        self.date = date.today()  #stores the corresponding test date
        self.startTime=0          #stores the timestamp when testing is starting for cat
        self.endTestTime=0        #stores the timestamp when testing is over for the cat
        self.group=group          #stores the cat's group number
        self.avg=0                #Initialize the average to be 0
        self.standardDev=0        #Initialize the standard deviation to be 0
        self.trialVariance=0      #Initialize the variance to be 0
        self.numOfTrials=0        #Counter for how many trials were performed for animal
        return
    
    ##########################################
    # Class Functions:
    # Getter Functions for animal's attributes
    ##########################################
    #Getter for the animal's name
    def getName(self):
        return self.name
    #Getter for the test date corresponding to the animal
    def getDate(self):
        return self.date
    #Getter for the animal's group number
    def getGroup(self):
        return self.group
    #Getter for the animal's number of trials completed
    def getNumOfTrials(self):
        return self.numOfTrials    
    #Getter for the average withdrawal time
    def getAvg(self):
        sample=self.time
        self.avg=mean(sample)
        return self.avg
    #Getter for time at specifed trial
    def getTimeAt(self,index):
        return self.time[index] 
    #Getter for the Standard Deviation of the animal's response times
    def getStdev(self):
        sample=self.time
        self.standardDev=pstdev(sample)
        return self.standardDev
    #Getter for the Variance of the animal's response times
    def getVar(self):
        sample=self.time
        self.trialVariance=pvariance(sample)
        return self.trialVariance

    ###################################################################################
    # Functions for turning on laser system and inserting withdrawal time into database
    ###################################################################################
    #Method for inserting a new withdrawl time into the list
    def insertTime(self,new_time):
        self.time.append(new_time)
        self.numOfTrials+=1
        return

    #Method for performing a trial and saving the data
    def trial(self):

        ############################################################################
        #Setup GPIO Board for RPI
       # GPIO.setmode(GPIO.BOARD)
       # GPIO.setup(13,GPIO.OUT)
       # GPIO.output(13,GPIO.LOW)
        ############################################################################

        #Run a trial and insert withdrawal time into a list for exporting
        t_1=animal.timer()

        ############################################################################
        #Clear the GPIO Pins, overide laser to be off
       # GPIO.setmode(GPIO.BOARD)
       # GPIO.setup(13,GPIO.OUT)
       # GPIO.output(13,GPIO.LOW)
        ############################################################################

        t_1=float(t_1) #Ensure the time is a valid time with decimals
        #Was the trial valid enough for the user?
        valid_trial=self.validInput(input("Keep trial? (Yes/No/Quit): "))
        if valid_trial==0:
            #perform a retrial
            self.trial()
        if valid_trial==1:        
            #When a valid trial is ran, 
            self.insertTime(t_1) #place withdrawal time into list
            if self.numOfTrials>1:
            #Update the analysis
                self.avg=self.getAvg()
                self.standardDev=self.getStdev()
                self.trialVariance=self.getVar()
            else:
                #If there is only 1 trial, dont update analysis
                self.avg=self.getTimeAt(0) #Average is the only valid trial
                self.standardDev=0
                self.trialVariance=0
        return
        
    #Method for recording withdrawal time
    def timer():
       # GPIO.setmode(GPIO.BOARD)
       # GPIO.setup(11,GPIO.IN)  #PhotoDiode Signal Pin
       # GPIO.setup(13,GPIO.OUT) #Output to the Laser
       # GPIO.output(13,GPIO.LOW) #Initialize to off
    
        print("\tProgram Started")
        
        #--------------------------------------------------------------
        
        t_1=input("") #REMOVE THIS LINE IN FUTURE THIS IS JUST FOR PROGRAMMING AT HOME
        
        print("\tPaw Placed time: %s seconds" % t_1)
        return t_1
        #--------------------------------------------------------------
        
        first_placed=False
        print("Searching for Paw")
        while first_placed is False:
            GPIO.output(13,GPIO.LOW)
            if GPIO.input(11)==0:
                first_placed=True
                GPIO.output(13,GPIO.HIGH) #Turn on the Laser
                print("Paw Placed")
                continue
        
        t_0=time()
        while True:
            if GPIO.input(11)==0:
                placed=True
               
            else:
                placed=False
                GPIO.output(13,0)

            if first_placed is True and placed is False:
                t_1=time()-t_0
                print("Paw Placed time: %s seconds" % t_1)
                GPIO.output(13,GPIO.LOW) #Turn off the laser
            
                break
        GPIO.output(13,GPIO.LOW)    #Possibly could comment out
        GPIO.cleanup()              #Could possibly comment out
        return t_1

    # Function for printing withdrawal times stored within list 
    def printTime(self):
        trialNum=1      #Counter for the current Trial
        #Print the animal's information and date
        print("Animal's Name:",self.name,"\tGroup:",self.group,"\tTest Date:",self.date,"\t Test Times:",self.startTime,"-",self.endTestTime)
        print("----------------------------------------------------------------------------")
        
        #Print the times for each trial
        for i in self.time:
            print("Trial",trialNum,":","%s seconds" % i)
            trialNum+=1
        #Print the animal's average time, standard deviation, variance
        print("\nAnalysis")
        
        #If there is only 1 trial, dont update analysis
        if self.numOfTrials<=1:
            self.avg=self.getTimeAt(0)  #Average will be the only valid trial
            self.standardDev=0          #Standard Dev and Variance would be 0 due to insufficient trials
            self.trialVariance=0
            print("***Not enough trial data to perform analysis***")
        #Otherwise if there are multiple valid trials, print analysis
        else:
            #Print the animal's analysis
            print("Average:", self.avg)
            print("Variance:", self.trialVariance)
            print("Standard Deviation:", self.standardDev,"\n")
        return 
    #Checks yes, no, or when to quit program
    def validInput(self,user_input):
        while True:
            if(user_input.lower()=="y") or user_input.lower()=="yes" or user_input=="1":
                output=1
                break
            if(user_input.lower()=="n") or user_input.lower()=="no" or user_input=="0":
                output=0
                break
            if(user_input.lower()=="q") or user_input.lower()=="quit" or user_input=="-1":
                output=-1
                break
            #Prompt for a new input if invalid
            user_input=input("Please Enter a Valid Input (Yes/No/Quit)")
        return output

#####################################################################
# Create a class to run a test and export all the animals information 
#####################################################################
class testAnimals:
    #No need for a constructor
    def __init__(self):
        self.numberOfGroups=0   #Keep track of the amount of groups avaliable
        self.numberOfAnimals=0  #Keep track of how many animals are tested upon
        return

    #####################################################
    # Functions for exporting results to spreadsheet
    #####################################################
    def exportingResults(self,results):
        resultTb=results
        workbookName=input("Save Output File As .xlsx?")
        #Check if the excel file exists, if not create it
        try: 
            #load the excel file with the corresponding workbook name
            wb1=openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        except FileNotFoundError:
            #Create the excel file with the corresponding workbook name
            wb1=self.exportSetup(workbookName)
            #Create enough sheets for each group
            for i in range(self.numberOfGroups+1):
                ws=wb1.create_sheet()
                ws.title="Group"+str(i)
                ws.append(["Group","Num of Group Trials","Group Avg","Animal Name","Trial Num","Withdrawal Times","Avg Time",
                            "Stdev","Trial Variance","Test Date", "Start Test Time","End Test Time"])
        
        #For 0 to max number of groups, load and write data from the results table into excel
        for i in range(self.numberOfGroups+1):
            curr_group=resultTb.arr[i]              #For readability, get the current group
            print(wb1.sheetnames)
            groupTimes=[]                           #Keeps track of all withdrawawl times in the group
            groupTrialNum=1                         #Counter for how many trials were within a group
            #check if in the correct group sheet
            if ("Group"+str(i)) in wb1.sheetnames:
                ws=wb1["Group"+str(i)]              #Get the correct group's sheet
                #Iterate across the list at the group index, where each element is a stored animal
                for j in range(len(curr_group)):
                    curr_animal=curr_group[j]       #For readability, get the current animal
                    trialNum=1                      #For printing the trial number corresponding to the withdrawal time
                    analyzeTime=[]                  #For performing analysis for all data up to a given point
                    for w_time in curr_animal.time:
                        analyzeTime.append(w_time)  #store time local to cat for analysis
                        groupTimes.append(w_time)   #store time local to group for analysis
                        #Print all the group and cat data to excel
                        ws.append([curr_animal.getGroup(),groupTrialNum,mean(groupTimes),curr_animal.getName(), trialNum, w_time, 
                                mean(analyzeTime), pstdev(analyzeTime), pvariance(analyzeTime)
                                ,curr_animal.date, curr_animal.startTime,curr_animal.endTestTime])
                        trialNum+=1                 
                        groupTrialNum+=1
            #Otherwise create as many sheets as needed for the number of groups avaliable
            else:
                ws=wb1.create_sheet()
                #Label the sheet based on the group number
                ws.title="Group"+str(i)
                #Create column labels
                ws.append(["Group","Num of Group Trials","Group Avg","Animal Name","Trial Num","Withdrawal Times","Avg Time",
                        "Stdev","Trial Variance","Test Date", "Start Test Time","End Test Time"])
        #Save the excel file
        wb1.save(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        # Close the workbook and return
        wb1.close()
        return

    #This function creates a spreadsheet based on the workbook name and labels the proper columns for the animal's data
    def exportSetup(self,workbookName):
        #Create a work book
        wb1=openpyxl.Workbook()
        #Create a worksheet within the workbook
        ws1=wb1.active
        ws1.title="Results"
        #initialize all the column names
        ws1.append(["Group","Num of Group Trials","Group Avg","Animal Name","Trial Num","Withdrawal Times","Avg Time",
                     "Stdev","Trial Variance","Test Date", "Start Test Time","End Test Time"])
        #prompt for an output file name
        wb1.save(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        #close the workbook
        wb1.close
        return wb1  #return the workbook so we can read/write to it with the other export functions
    
    ###############################################################################             
    #  Functions for starting the test and prompting for an animals information                          
    ###############################################################################    
    #Prompts the user for name and number to construct animal
    def promptAnimalInfo():
        #Prompt for name of Animal
        name=input("Enter subject's Name: ")
        #Prompt for a valid integer group number
        while True:
            try:
                group = int(input("Enter subject's Group number, minimum of 1: "))
                break
            except ValueError:
                print("Please enter Group number as an integer")  
            continue
        #Return Current Animal's name and number
        return name,group
    
    #Checks yes, no, or when to quit program
    def validInput(self,user_input):
        while True:
            if(user_input.lower()=="y") or user_input.lower()=="yes" or user_input=="1":
                output=1
                break
            if(user_input.lower()=="n") or user_input.lower()=="no" or user_input=="0":
                output=0
                break
            if(user_input.lower()=="q") or user_input.lower()=="quit" or user_input=="-1":
                output=-1
                break
            #prompt for a new input if invalid
            user_input=input("Please Enter a Valid Input (Yes/No/Quit)")
        return output

    #This method is what starts the entire system to even run a test
    def startExperiment(self):
        #Create a table to hold any results
        results=resultTb()

       #Prompt user to start program
        print("\n*******Welcome to the Plantar Thermal Laser test.*******")
        beginProgram=self.validInput(input("Would you like to run a trial? (Yes/No/Quit): "))

        #Determine valid response, if so create a animal as an object
        while (beginProgram==1):
            #Create an animal data structure
            name,group=testAnimals.promptAnimalInfo()
            curr_animal=animal(name,group)
            curr_animal.startTime=datetime.now().strftime("%H:%M:%S")  #get the timestamp when testing is beginning
            trialNum=1                                                  #Keeps track of trial of current animal

            #keep track of the maximum amount of groups
            if group>self.numberOfGroups:
                self.numberOfGroups=group 

            #Begin Testing a animal
            runTrial=1      #Flag to determine which animal is being tested on
            #Run as many trials as wanted for this animal
            while (runTrial==1):
                #Perform a Trial, get withdrawal time
                print("Trial ",trialNum,":")
                curr_animal.trial()

                #Prompt to perform another trial
                nextTrial=self.validInput(input("Run another trial for this animal?(Yes/No/Quit) "))
                if (nextTrial==1): #Keep testing current animal
                    trialNum+=1
                    runTrial=1 
                if (nextTrial==0): #Stop testing current animal
                    trialNum=0
                    runTrial=0
                if(nextTrial==-1): #Quit the program
                    beginProgram==-1
                    return

            #Print animal's information after testing 
            curr_animal.endTestTime=datetime.now().strftime("%H:%M:%S")  #get the timestamp when testing is over for cat
            curr_animal.printTime()

            #Insert the animal into the table based upon their group number
            results.insert(curr_animal)

            #Ask user to test another animal
            testAnotheranimal=self.validInput(input("\nTest a new animal?(Yes/No/Quit)"))
            #If the User Answers Yes
            if (testAnotheranimal==1): 
                beginProgram=1 #Jump back to start of loop and create new animal

            #If the User Answers No, print and export results, then exit the program
            if ((testAnotheranimal==0) or (testAnotheranimal==-1)): 
                results.printAll()
                self.exportingResults(results)
                beginProgram=0 #Exit loop
                        
        #Exit program when testing is stopped
        if (beginProgram==0 ) or (beginProgram==-1):     
            print("*****Ending Testing*****")
        return

#########################################
#           Driver Function             #
#########################################
def main():
    #create a test and start the experiment
    test=testAnimals()
    test.startExperiment()
    #exit program
    return 0

##Main definition##
if __name__ == "__main__":
    main()    
