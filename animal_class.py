###########################################
## 
## Animal Class for thermal laser
## SD Team 10: Robert Bara, Zari Grandy, Ezra Galapo, Tyiana Smith
## 
## Date:    3/10/2022
## Version: 2.3
##
## Description:
##  This python program allows for the thermal laser plantar test to run by prompting the user if
##  they would like to perform a test. The class creates an animal, performs a laser test, stores the withdrawal
##  times, as well as performs analysis such as standard dev, average, and variance calculations
##
## Usage:
##  Running: python3 animal_class.py (run it here, eventually build a make file-Rob)
##
########################################

############ Includes ##################
from datetime import date   
from datetime import datetime
import os
from os import stat
from re import S
from sre_constants import MAXGROUPS             # Import date class from datetime module
from statistics import pstdev    # Import statistics to utilize mean, standard deviation, and variance calculations
from statistics import pvariance
from statistics import mean
from enum import Enum
import RPi.GPIO as GPIO        # Import General-Purpose In/Out for RPI4 to control laser and photodiode
from time import time
from numpy import save           # Import time to calculate withdrawal time
import openpyxl                 # Import for exporting the animal's results to spreadsheets
############ animal Class #################
class animal:
    ##############################
    #Constructor for animal
    ##############################
    def __init__(self,name,group):
        self.name=name            #animal is named from user input
        self.time=[]              #create a list of all the animal's withdrawal times
        self.date = date.today()  #stores the corresponding test date
        self.startTime=0 #date.strftime("%H:%M:%S")  #stores the timestamp when testing is starting for cat
        self.endTestTime=0        #stores the timestamp when testing is over for the cat
        self.group=group          #stores the cat's group number
        self.avg=0                #Initialize the average to be 0
        self.standardDev=0        #Initialize the standard deviation to be 0
        self.trialVariance=0      #Initialize the variance to be 0
        self.numOfTrials=0
        return
    
    ##########################################
    # Class Functions:
    # Getter Functions for animal's attributes
    ##########################################
    #Getter for the animal's name
    def getName(self):
        return self.name
    #Getter for the test date corresponding to the animal
    def getDate(self):
        return self.date
    #Getter for the animal's group number
    def getGroup(self):
        return self.group
    #Getter for the animal's number of trials completed
    def getNumOfTrials(self):
        return self.numOfTrials    
    #Getter for the average withdrawal time
    def getAvg(self):
        sample=self.time
        self.avg=mean(sample)
        return self.avg
    #Getter for time at specifed trial
    def getTimeAt(self,index):
        return self.time[index] 
    #Getter for the Standard Deviation of the animal's response times
    def getStdev(self):
        sample=self.time
        self.standardDev=pstdev(sample)
        return self.standardDev
    #Getter for the Variance of the animal's response times
    def getVar(self):
        sample=self.time
        self.trialVariance=pvariance(sample)
        return self.trialVariance

    ###################################################################################
    # Functions for turning on laser system and inserting withdrawal time into database
    ###################################################################################
    #Method for inserting a new withdrawl time into the list
    def insertTime(self,new_time):
        self.time.append(new_time)
        self.numOfTrials+=1
        return

    #Method for performing a trial and saving the data
    def trial(self):

        ############################################################################
        #Setup GPIO Board for RPI
        GPIO.setmode(GPIO.BOARD)
        GPIO.setup(13,GPIO.OUT)
        GPIO.output(13,GPIO.LOW)
        ############################################################################

        #Run a trial and insert withdrawal time into a list for exporting
        t_1=animal.timer()

        ############################################################################
        #Clear the GPIO Pins, overide laser to be off
        GPIO.setmode(GPIO.BOARD)
        GPIO.setup(13,GPIO.OUT)
        GPIO.output(13,GPIO.LOW)
        ############################################################################

        t_1=float(t_1) #Ensure the time is a valid time with decimals
        #Was the trial valid enough for the user?
        valid_trial=input("Keep trial? (Yes/No): ")
        if valid_trial != "y":
            #perform a retrial
            self.trial()
        else:        
            #When a valid trial is ran, 
            self.insertTime(t_1) #place withdrawal time into list
            if self.numOfTrials>1:
            #Update the analysis
                self.avg=self.getAvg()
                self.standardDev=self.getStdev()
                self.trialVariance=self.getVar()
            else:
                #If there is only 1 trial, dont update analysis
                self.avg=self.getTimeAt(0) #Average is the only valid trial
                self.standardDev=0
                self.trialVariance=0

        return
        
    #Method for recording withdrawal time
    def timer():
        GPIO.setmode(GPIO.BOARD)
        GPIO.setup(11,GPIO.IN)  #PhotoDiode Signal Pin
        GPIO.setup(13,GPIO.OUT) #Output to the Laser
        GPIO.output(13,GPIO.LOW) #Initialize to off
    
        print("\tProgram Started")
        
        #--------------------------------------------------------------
        
        #t_1=input("") #REMOVE THIS LINE IN FUTURE THIS IS JUST FOR PROGRAMMING AT HOME
        
        #print("\tPaw Placed time: %s seconds" % t_1)
        #return t_1
        #--------------------------------------------------------------
        
        first_placed=False
        print("Searching for Paw")
        while first_placed is False:
            GPIO.output(13,GPIO.LOW)
            if GPIO.input(11)==0:
                first_placed=True
                GPIO.output(13,GPIO.HIGH) #Turn on the Laser
                print("Paw Placed")
                continue
        
        t_0=time()
        while True:
            if GPIO.input(11)==0:
                placed=True
               
            else:
                placed=False
                GPIO.output(13,0)

            if first_placed is True and placed is False:
                t_1=time()-t_0
                print("Paw Placed time: %s seconds" % t_1)
                GPIO.output(13,GPIO.LOW) #Turn off the laser
            
                break
        GPIO.output(13,GPIO.LOW)    #Possibly could comment out
        GPIO.cleanup()              #Could possibly comment out
        return t_1

    # Function for printing withdrawal times stored within list 
    def printTime(self):
        trialNum=1      #Counter for the current Trial
        #Print the animal's information and date
        print("Animal's Name:",self.name,"\tGroup:",self.group,"\tTest Date:",self.date,"\t Test Times:",self.startTime,"-",self.endTestTime)
        print("----------------------------------------------------------------------------")
        
        #Print the times for each trial
        for i in self.time:
            print("Trial",trialNum,":","%s seconds" % i)
            trialNum+=1
        #Print the animal's average time, standard deviation, variance
        print("\nAnalysis")
        
        #If there is only 1 trial, dont update analysis
        if self.numOfTrials<=1:
            self.avg=self.getTimeAt(0)  #Average will be the only valid trial
            self.standardDev=0          #Standard Dev and Variance would be 0 due to insufficient trials
            self.trialVariance=0
            print("***Not enough trial data to perform analysis***")
        #Otherwise if there are multiple valid trials, print analysis
        else:
            #Print the animal's analysis
            print("Average:", self.avg)
            print("Variance:", self.trialVariance)
            print("Standard Deviation:", self.standardDev,"\n")
        return 
    

#####################################################################
# Create a class to run a test and store all the animals information 
#####################################################################
class testAnimals:
    #No need for a constructor
    def __init__(self) -> None:
        pass
    ############################################################
    # Functions to Print results in various ways to the terminal
    ############################################################
    #Print the results of the Passed in Group number
    def printGroup(self, saveAnimals, group):
        print("\n--------------------------------------")
        print("Group",group,"Results:")
        print("--------------------------------------\n")
        #If the group's number matches the group argument, print Animal's information
        for i in saveAnimals:
            if(i.group==group): 
                i.printTime() 
        return

    #Prints the results of all groups in orderS
    def printAllGroups(self, saveAnimals):
        numOfGroups=self.maxGroup(saveAnimals)
        for i in range(numOfGroups+1):
            self.printGroup(saveAnimals,i)
        return
    #Function determines how many groups are there    
    def maxGroup(self,saveAnimals):
        numOfGroups=0
        for i in saveAnimals:
            if(i.group>numOfGroups):
                numOfGroups=i.group
        return numOfGroups

    #This function prints today's Test Animal's results
    def printTodaysResults(self, saveAnimals):
        print("\n--------------------------------------")
        print("Test Results:")
        print("--------------------------------------\n")
        #Print every Animal's information
        for i in saveAnimals:
            i.printTime() 
        return
    #####################################################
    # Functions for exporting results to spreadsheet
    #####################################################
    #This function creates a spreadsheet and labels the proper columns for the animal's data
    def exportSetup(self):
        #Create a work book
        wb1=openpyxl.Workbook()
        #Create a worksheet within the workbook
        ws1=wb1.active
        ws1.title="Group"
        #initialize all the column names
        ws1.append(["Group","Number of Groups","Group Avg","Animal Name","Trial Num","Withdrawal Times","Avg Time", "Stdev","Trial Variance"])
        #prompt for an output file name
        workbookName=input("Save Output File As .xlsx?")
        wb1.save(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        #close the workbook
        wb1.close
        return workbookName #return the workbook name so we can write to it with the other export functions

    #This function exports the animals that were tested on the current day, to a spreadsheet
    def exportResults(self,saveAnimals,saveAnimalsLength):
        #Call excel setup function
        workbookName=self.exportSetup()
        #Load the workbook that has been created
        wb1=openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        #Write Data to the current sheet
        ws1=wb1.active
        exportAnimals=saveAnimals
        #Loop through the saved animals
        for i in range(saveAnimalsLength):
            #remove the animal at the front of the list 
            currAnimal=exportAnimals.pop(0)
            #set the counter for the trial number
            trialNum=1
            #Loop through the Current Animal's withdrawal times
            for j in currAnimal.time:
                withdrawalTime=j
                group=currAnimal.getGroup()
                #Print the animal's attributes
                ws1.append([group,"n/a","n/a",currAnimal.getName(), trialNum, withdrawalTime, 
                            currAnimal.avg, currAnimal.standardDev, currAnimal.trialVariance])       
                trialNum+=1
        # When exporting is done, save the workbook
        wb1.save(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        # Close the workbook and return
        wb1.close()
        return

    #Function for printing every group in order (Not finished yet)
    def exportAllGroups(self, saveAnimals):
        #Call excel setup function
        workbookName=self.exportSetup()
        #Load the workbook that has been created
        wb1=openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        #Write Data to the current sheet
        ws1=wb1.active
        exportAnimals=saveAnimals
        #Loop through the saved animals by group
        numOfGroups=self.maxGroup(exportAnimals)
        for k in range(numOfGroups+1):
            for i in exportAnimals:
                if(i.group==k): 
                    #remove the animal at the front of the list 
                    currAnimal=exportAnimals.pop(0)
                    #set the counter for the trial number
                    trialNum=1
                    #Loop through the Current Animal's withdrawal times
                    for j in currAnimal.time:
                        withdrawalTime=j
                        #Print the animal's attributes
                        ws1.append(["n/a","n/a","n/a",currAnimal.getName(), trialNum, withdrawalTime, 
                                    currAnimal.avg, currAnimal.standardDev, currAnimal.trialVariance])       
                        trialNum+=1
        # When exporting is done, save the workbook
        wb1.save(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        # Close the workbook and return
        wb1.close()
        return
    
    ###############################################################################             
    #  Functions for starting the test and prompting for an animals information                          
    ###############################################################################    
    #Prompts the user for name and number to construct animal
    def promptAnimalInfo():
        #Create a animal with their information
        #Prompt for name of Animal
        name=input("Enter subject's Name: ")
        #Prompt for a valid integer group number
        while True:
            try:
                group = int(input("Enter subject's Group number: "))
                break
            except ValueError:
                print("Please enter Group number as an integer")  
            continue
        #Return Current Animal's name and number
        return name,group
    
    #This method is what starts the entire system to even run a test
    def startExperiment(self):
        #GPIO.setup(4,GPIO.OUT) #PIN OF LASER
        #GPIO.output(4,GPIO.LOW) #initialize to low
        
        saveAnimals=[]  #Create a list to store all animal's information
        saveAnimalsLength=0 #Keep track of the size of the list
       #Prompt user to start program
        print("\n*******Welcome to the Plantar Thermal Laser test.*******")
        beginProgram=input("Would you like to run a trial? (Yes/No): ")
        print(beginProgram)

        #Determine valid response, if so create a animal as an object
        while beginProgram == "y":
            #Create an animal data structure
            name,group=testAnimals.promptAnimalInfo()
            curr_animal=animal(name,group)
            #Begin Testing a animal
            trialNum=1 #Keeps track of trial of current animal
            runTrial=1 #Flag to determine which animal is being tested on
            #Run as many trials as wanted for this animal
            while (runTrial==1):
                #Perform a Trial, get withdrawal time
                print("Trial ",trialNum,":")
                curr_animal.trial()

                #Prompt to perform another trial
                nextTrial=input("Run another trial for this animal?(Yes/No) ")
                if nextTrial == "y": #Keep testing current animal
                    trialNum+=1
                    runTrial=1 
                if nextTrial == "n": #Stop testing current animal
                    trialNum=0
                    runTrial=0
        #Print animal's information after testing 
            #curr_animal.endTestTime=date.strftime("%H:%M:%S")  #get the timestamp when testing is over for cat
            curr_animal.printTime()
        
            #Ask user to test another animal
            testAnotheranimal=input("\nTest a new animal?(Yes/No)")
            #If the User Answers Yes
            if testAnotheranimal == "y": 
                saveAnimals.append(curr_animal) #add another animal to the list
                saveAnimalsLength+=1            #update the length of the list
                beginProgram="y" #Jump back to start of loop and create new animal
            #If the User Answers No
            if testAnotheranimal == "n": 
                saveAnimals.append(curr_animal) #add another animal to the list
                saveAnimalsLength+=1            #update the length of the list
                self.printTodaysResults(saveAnimals)
                
                #Testing printing per group
                #self.printGroup(saveAnimals,0)
                #self.printGroup(saveAnimals,1)
                #self.printAllGroups(saveAnimals)
                self.exportResults(saveAnimals,saveAnimalsLength)
                #self.exportAllGroups(saveAnimals)
                beginProgram="n" #Exit loop
                        
        #Exit program when testing is stopped
        if beginProgram == "n":     
            print("*****Ending Testing*****")
        return

#########################################
#           Driver Function             #
#########################################
def main():
    #create a test and start the experiment
    test=testAnimals()
    test.startExperiment()
    #exit program
    return 0

##Main definition##
if __name__ == "__main__":
    main()    
