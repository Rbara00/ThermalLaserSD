###########################################
## 
## Animal Class for thermal laser
## SD Team 10: Robert Bara, Zari Grandy, Ezra Galapo, Tyiana Smith
## 
## Author: Robert Bara
## Date:    9/6/2022
## Version: 1.0
##
## Description:
##  This python file is responsible for taking data from the laser test and exporting it to excel. Additionally,
##  the results automatically generate a bar graph comparing the group averages and standard deviation
##
## Usage:
##  This program is invoked by test_class.py
##
########################################
# Imports
from tkinter import filedialog
import os
from statistics import pstdev, pvariance, mean   # Import statistics to utilize mean, standard deviation, and variance calculations
import openpyxl      # Import for exporting the animal's results to spreadsheets
import openpyxl.chart

#####################################################
# Functions for exporting results to spreadsheet
#####################################################
class export:
    #This function creates a spreadsheet based on the workbook name and labels the proper columns for the animal's data
    def exportSetup(self,workbookName):
        #Create a work book
        wb1=openpyxl.Workbook()
        #prompt for an output file name
        wb1.save(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        #close the workbook
        wb1.close
        return wb1  #return the workbook so we can read/write to it with the other export functions
    
    #Prompts the user to load/save an excel file
    def load(self):
        filename = filedialog.asksaveasfilename(initialdir=os.path.join(os.path.dirname(os.path.abspath(__file__))),
                                        defaultextension='.xlsx',
                                        filetypes=[
                                            ("xlsx",".xlsx"),
                                            ("xls", ".xls"),
                                            ('xl','.xl'),
                                            ('sm','.sm'),
                                            ('xtlx','.xtlx'),
                                            ('xltm','.xltm'),
                                            ("All files", ".*"),])
        return filename

    
    #This function checks for the appropriate test result spreadsheet and exports/analyzes any new test results to the sheet
    def exportResults(self,results, numOfGroups):
        #Check if the excel file exists, if not create it
        try: 
            workbookName=self.load()
            wb1=openpyxl.load_workbook(workbookName)
        except FileNotFoundError:
            #Create the excel file with the corresponding workbook name
            wb1=self.exportSetup(workbookName)
            #Create enough sheets for each group
            for i in range(1,numOfGroups+1):
                ws=wb1.create_sheet()
                ws.title="Group"+str(i)
                self.exportLabels(ws)       #write labels to excel sheet
                ws=self.columnWidth(ws)     #Format the Column Width

        #Write the test's results data to excel and compute analysis
        wb1=self.exportDataToResults(wb1, results, numOfGroups)

        #Save the excel file
        wb1.save(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        # Close the workbook and return
        wb1.close()
        return workbookName

    #This function writes the test results to the correct group's sheet, or creates a blank sheet for future testing
    def exportDataToResults(self, wb1, resultTb,numOfGroups):
        #For 1 to max number of groups, load and write data from the results table into excel
        for i in range(1,numOfGroups+1):
            curr_group=resultTb.arr[i]              #For readability, get the current group
            
            #check if in the correct group sheet
            if ("Group"+str(i)) in wb1.sheetnames:
                ws=wb1["Group"+str(i)]              #Get the correct group's sheet
                max_row=ws.max_row                  #Get the max row of the sheet
                groupTrialNum=max_row    #update the total number of trials per group
            
                #Iterate across the list at the group index, where each element is a stored animal
                for j in range(len(curr_group)):
                    curr_animal=curr_group[j]       #For readability, get the current animal
                    trialNum=1                      #For printing the trial number corresponding to the withdrawal time
                    analyzeTime=[]                  #For performing analysis for all data up to a given point
                   
                    #Export and Analyze each withdrawal time for the current animal
                    for w_time in curr_animal.time:
                        analyzeTime.append(w_time)  #store current time local to animal for analysis
                        groupAvg='=AVERAGE(G2:G'+str(groupTrialNum+1)+')'
                        groupStdev='=STDEVP(G2:G'+str(groupTrialNum+1)+')'
                        #Print all the group and animals data to excel
                        ws.append([curr_animal.getGroup(), groupTrialNum, groupAvg, groupStdev, 
                                curr_animal.getName(), trialNum, w_time, mean(analyzeTime), 
                                pstdev(analyzeTime), pvariance(analyzeTime), curr_animal.date, 
                                curr_animal.startTime, curr_animal.endTestTime])
                        #Increase trial number counters
                        trialNum+=1                 
                        groupTrialNum+=1

            #Otherwise create as many sheets as needed for the number of groups avaliable
            else:
                ws=wb1.create_sheet()
                #Label the sheet based on the group number
                ws.title="Group"+str(i)
                #Create column labels
                self.exportLabels(ws)
                #Write data to new sheets through recursion
                self.exportDataToResults(wb1,resultTb)  
        #Return updated workbook
        return wb1

    #This function creates a sheet specifically to graph the results of all group's average times and standard deviation
    def graphResults(self,workbookName):
        #Check if the excel file exists, if not return
        try: 
            #load the excel file with the corresponding workbook name
            wb1=openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        except FileNotFoundError:
            print("Invalid file name")
            return

        #Remove the results sheet and Re-Analyze
        if ("Results") in wb1.sheetnames:
            wb1.remove(wb1["Results"])
        #Create the graph results sheet
        ws0=wb1.create_sheet()
        ws0.title="Results"
        ws0.append(["Group","Avg Times","Stdev"])

        #Obtain the latest Group Averages and standard deviation
        ws1=wb1["Group"+str(1)]
        max_sheet=int(len(wb1.sheetnames))
        for i in range(1,max_sheet):
            #check if in the correct group sheet
            if ("Group"+str(i)) in wb1.sheetnames:
                ws1=wb1["Group"+str(i)]
                max_row=ws1.max_row
                #Check if there is data to graph
                if isinstance(ws1.cell(max_row,1).value,str) is False:
                    #Locate the current group's latest group average and standard deviation
                    groupNum='=Group'+str(i)+'!A'+str(max_row)
                    groupAvg='=Group'+str(i)+'!C'+str(max_row)
                    groupStdev='=Group'+str(i)+'!D'+str(max_row)
                    ws0.append([groupNum,groupAvg,groupStdev])     #write data to the results sheet
        
        #Create the bar graph from the results    
        ws0=self.createBarGraphResults(ws0)

        #Save and close the excel file
        wb1.save(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        wb1.close()
        return
    
    #This function generates a bar graph for the results sheet
    def createBarGraphResults(self,ws0,):
        #Generate a Bar Graph for averages
        chart1 = openpyxl.chart.bar_chart.BarChart()
        chart1.type = "col"
        chart1.style = 3
        chart1.title = "Average Responses of All Groups"
        chart1.y_axis.title = 'Avg and Stdev of Withdrawal Times'
        chart1.x_axis.title = 'Group Numbers'

        #Data to be graphed
        data = openpyxl.chart.Reference(ws0, min_col=2, min_row=1, max_row=ws0.max_row, max_col=ws0.max_column)
        cats = openpyxl.chart.Reference(ws0, min_col=1, min_row=2, max_row=ws0.max_row)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 8
        ws0.add_chart(chart1, "G2")     #Position where graph will be added

        #Return the updated results sheet with graph
        return ws0

    #Formats the excel column width nicely
    def columnWidth(self,ws):
        ws.column_dimensions['A'].width=7       #Group Number
        ws.column_dimensions['B'].width=10       #Number of trials in group
        ws.column_dimensions['C'].width=12      #Group's Average Response Time
        ws.column_dimensions['D'].width=12      #Group's Standard Deviation
        ws.column_dimensions['E'].width=12      #Name of Animal
        ws.column_dimensions['F'].width=9       #Number of trials per Animal
        ws.column_dimensions['G'].width=14      #Animal Response time at given trial
        ws.column_dimensions['H'].width=9       #Animal's Average Response Time
        ws.column_dimensions['I'].width=9       #Animal's Standard Deviation
        ws.column_dimensions['J'].width=9      #Animal's Variance
        ws.column_dimensions['K'].width=12      #Date of Testing
        ws.column_dimensions['L'].width=11      #Starting time for testing animal
        ws.column_dimensions['M'].width=11       #Ending time for testing animal
        return ws
    
    #Writes the labels for a group's sheet
    def exportLabels(self,ws):
        ws.append(["Group","Group Trial Num","Group Avg","Group Stdev","Animal Name",
                   "Trial Num","Response Times","Avg Time","Stdev","Trial Var",
                   "Test Date", "Start Test Time","End Test Time"])
        return ws