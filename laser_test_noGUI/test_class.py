#!/usr/bin/python3
###########################################
## 
## Test Class for thermal laser
## SD Team 10: Robert Bara, Zari Grandy, Ezra Galapo, Tyiana Smith
## 
## Author: Robert Bara
## Date:    9/6/2022
## Version: 1.3
##
## Description:
##  This Python script invokes the other classes to provide a structure for the laser test and
##  group analysis. This version of the script is for a test that will not need a GUI, and therefore
##  is ran within the terminal.
##
########################################
from table_class import*
from animal_class import*
from datetime import datetime
import os
from statistics import geometric_mean, pstdev, pvariance, mean   # Import statistics to utilize mean, standard deviation, and variance calculations
from math import sqrt
#import RPi.GPIO as GPIO        # Import General-Purpose In/Out for RPI4 to control laser and photodiode
from time import time
import openpyxl                 # Import for exporting the animal's results to spreadsheets
import openpyxl.chart
#####################################################################
# Create a class to run a test and export all the animals information 
#####################################################################
class testAnimals:
    #No need for a constructor
    def __init__(self):
        self.numberOfGroups=0   #Keep track of the amount of groups avaliable
        self.numberOfAnimals=0  #Keep track of how many animals are tested upon
        return

    #####################################################
    # Functions for exporting results to spreadsheet
    #####################################################

    #This function creates a spreadsheet based on the workbook name and labels the proper columns for the animal's data
    def exportSetup(self,workbookName):
        #Create a work book
        wb1=openpyxl.Workbook()
        #prompt for an output file name
        wb1.save(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        #close the workbook
        wb1.close
        return wb1  #return the workbook so we can read/write to it with the other export functions
    
    #This function checks for the appropriate test result spreadsheet and exports/analyzes any new test results to the sheet
    def exportingResults(self,results):
        resultTb=results
        #Prompt to save results as excel file
        workbookName=input("Saving output to .xlsx file. Enter Filename: ")+str('.xlsx')
        #Check if the excel file exists, if not create it
        try: 
            #load the excel file with the corresponding workbook name
            wb1=openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)),(workbookName)))
        except FileNotFoundError:
            #Create the excel file with the corresponding workbook name
            wb1=self.exportSetup(workbookName)
            #Create enough sheets for each group
            for i in range(1,self.numberOfGroups+1):
                ws=wb1.create_sheet()
                ws.title="Group"+str(i)
                #write labels to excel sheet
                self.exportLabels(ws)
                ws=self.columnWidth(ws)     #Format the Column Width

        #Write the test's results data to excel and compute analysis
        wb1=self.exportDataToResults(wb1, resultTb)

        #Save the excel file
        wb1.save(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        # Close the workbook and return
        wb1.close()
        return workbookName

    #This function writes the test results to the correct group's sheet, or creates a blank sheet for future testing
    def exportDataToResults(self, wb1, resultTb):
        #For 1 to max number of groups, load and write data from the results table into excel
        for i in range(1,self.numberOfGroups+1):
            curr_group=resultTb.arr[i]              #For readability, get the current group
            
            #check if in the correct group sheet
            if ("Group"+str(i)) in wb1.sheetnames:
                ws=wb1["Group"+str(i)]              #Get the correct group's sheet
                max_row=ws.max_row                  #Get the max row of the sheet
                groupTrialNum=max_row    #update the total number of trials per group
            
                #Iterate across the list at the group index, where each element is a stored animal
                for j in range(len(curr_group)):
                    curr_animal=curr_group[j]       #For readability, get the current animal
                    trialNum=1                      #For printing the trial number corresponding to the withdrawal time
                    analyzeTime=[]                  #For performing analysis for all data up to a given point
                   
                    #Export and Analyze each withdrawal time for the current animal
                    for w_time in curr_animal.time:
                        analyzeTime.append(w_time)  #store current time local to animal for analysis
                        groupAvg='=AVERAGE(G2:G'+str(groupTrialNum+1)+')'
                        groupStdev='=STDEVP(G2:G'+str(groupTrialNum+1)+')'
                        #Print all the group and animals data to excel
                        ws.append([curr_animal.getGroup(), groupTrialNum, groupAvg, groupStdev, 
                                curr_animal.getName(), trialNum, w_time, mean(analyzeTime), 
                                pstdev(analyzeTime), pvariance(analyzeTime), curr_animal.date, 
                                curr_animal.startTime, curr_animal.endTestTime])
                        #Increase trial number counters
                        trialNum+=1                 
                        groupTrialNum+=1

            #Otherwise create as many sheets as needed for the number of groups avaliable
            else:
                ws=wb1.create_sheet()
                #Label the sheet based on the group number
                ws.title="Group"+str(i)
                #Create column labels
                self.exportLabels(ws)
                #Write data to new sheets through recursion
                self.exportDataToResults(wb1,resultTb)  
        #Return updated workbook
        return wb1

    #This function creates a sheet specifically to graph the results of all group's average times and standard deviation
    def graphResults(self,workbookName):
        #Check if the excel file exists, if not return
        try: 
            #load the excel file with the corresponding workbook name
            wb1=openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        except FileNotFoundError:
            print("Invalid file name")
            return

        #Remove the results sheet and Re-Analyze
        if ("Results") in wb1.sheetnames:
            wb1.remove(wb1["Results"])
        #Create the graph results sheet
        ws0=wb1.create_sheet()
        ws0.title="Results"
        ws0.append(["Group","Avg Times","Stdev"])

        #Obtain the latest Group Averages and standard deviation
        ws1=wb1["Group"+str(1)]
        max_sheet=int(len(wb1.sheetnames))
        for i in range(1,max_sheet):
            #check if in the correct group sheet
            if ("Group"+str(i)) in wb1.sheetnames:
                ws1=wb1["Group"+str(i)]
                max_row=ws1.max_row
                #Check if there is data to graph
                if isinstance(ws1.cell(max_row,1).value,str) is False:
                    #Locate the current group's latest group average and standard deviation
                    groupNum='=Group'+str(i)+'!A'+str(max_row)
                    groupAvg='=Group'+str(i)+'!C'+str(max_row)
                    groupStdev='=Group'+str(i)+'!D'+str(max_row)
                    ws0.append([groupNum,groupAvg,groupStdev])     #write data to the results sheet
        
        #Create the bar graph from the results    
        ws0=self.createBarGraphResults(ws0)

        #Save and close the excel file
        wb1.save(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        wb1.close()
        return
    
    #This function generates a bar graph for the results sheet
    def createBarGraphResults(self,ws0,):
        #Generate a Bar Graph for averages
        chart1 = openpyxl.chart.bar_chart.BarChart()
        chart1.type = "col"
        chart1.style = 3
        chart1.title = "Average Responses of All Groups"
        chart1.y_axis.title = 'Avg and Stdev of Withdrawal Times'
        chart1.x_axis.title = 'Group Numbers'

        #Data to be graphed
        data = openpyxl.chart.Reference(ws0, min_col=2, min_row=1, max_row=ws0.max_row, max_col=ws0.max_column)
        cats = openpyxl.chart.Reference(ws0, min_col=1, min_row=2, max_row=ws0.max_row)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 8
        ws0.add_chart(chart1, "G2")     #Position where graph will be added

        #Return the updated results sheet with graph
        return ws0

    #Formats the excel column width nicely
    def columnWidth(self,ws):
        ws.column_dimensions['A'].width=7       #Group Number
        ws.column_dimensions['B'].width=10       #Number of trials in group
        ws.column_dimensions['C'].width=12      #Group's Average Response Time
        ws.column_dimensions['D'].width=12      #Group's Standard Deviation
        ws.column_dimensions['E'].width=12      #Name of Animal
        ws.column_dimensions['F'].width=9       #Number of trials per Animal
        ws.column_dimensions['G'].width=14      #Animal Response time at given trial
        ws.column_dimensions['H'].width=9       #Animal's Average Response Time
        ws.column_dimensions['I'].width=9       #Animal's Standard Deviation
        ws.column_dimensions['J'].width=9      #Animal's Variance
        ws.column_dimensions['K'].width=12      #Date of Testing
        ws.column_dimensions['L'].width=11      #Starting time for testing animal
        ws.column_dimensions['M'].width=11       #Ending time for testing animal
        return ws
    
    #Writes the labels for a group's sheet
    def exportLabels(self,ws):
        ws.append(["Group","Group Trial Num","Group Avg","Group Stdev","Animal Name",
                   "Trial Num","Response Times","Avg Time","Stdev","Trial Var",
                   "Test Date", "Start Test Time","End Test Time"])
        return ws

    ###############################################################################             
    #  Functions for starting the test and prompting for an animals information                          
    ###############################################################################    
    #Prompts the user for name and number to construct animal
    def promptAnimalInfo():
        #Prompt for name of Animal
        name=input("Enter subject's Name: ")
        #Prompt for a valid integer group number
        while True:
            try:
                group = int(input("Enter subject's Group number, minimum of 1: "))
                if group>=1:
                    break
            except ValueError:
                print("Please enter Group number as an integer")  
            continue
        #Return Current Animal's name and number
        return name,group
    
    #Checks yes, no, or when to quit program
    def validInput(self,user_input):
        while True:
            if(user_input.lower()=="y") or user_input.lower()=="yes" or user_input=="1":
                output=1
                break
            if(user_input.lower()=="n") or user_input.lower()=="no" or user_input=="0":
                output=0
                break
            if(user_input.lower()=="q") or user_input.lower()=="quit" or user_input=="-1":
                output=-1
                break
            #prompt for a new input if invalid
            user_input=input("Please Enter a Valid Input (Yes/No/Quit)")
        return output

    #This method is what starts the entire system to even run a test
    def startExperiment(self):
        #Create a table to hold any results
        results=resultTb()

       #Prompt user to start program
        print("\n*******Welcome to the Thermal Plantar Laser test.*******")
        beginProgram=self.validInput(input("Would you like to run a trial? (Yes/No/Quit): "))

        #Determine valid response, if so create a animal as an object
        while (beginProgram==1):
            #Create an animal data structure
            name,group=testAnimals.promptAnimalInfo()
            curr_animal=animal(name,group)
            curr_animal.startTime=datetime.now().strftime("%H:%M:%S")  #get the timestamp when testing is beginning
            trialNum=1                                                  #Keeps track of trial of current animal

            #keep track of the maximum amount of groups
            if group>self.numberOfGroups:
                self.numberOfGroups=group 

            #Begin Testing a animal
            runTrial=1      #Flag to determine which animal is being tested on
            #Run as many trials as wanted for this animal
            while (runTrial==1):
                #Perform a Trial, get withdrawal time
                print("Trial ",trialNum,":")
                curr_animal.trial()

                #Prompt to perform another trial
                nextTrial=self.validInput(input("Run another trial for this animal?(Yes/No/Quit) "))
                if (nextTrial==1): #Keep testing current animal
                    trialNum+=1
                    runTrial=1 
                if (nextTrial<=0): #Stop testing current animal
                    trialNum=0
                    runTrial=0

            #Print animal's information after testing 
            curr_animal.endTestTime=datetime.now().strftime("%H:%M:%S")  #get the timestamp when testing is over for cat
            curr_animal.printTime()

            #Insert the animal into the table based upon their group number
            results.insert(curr_animal)

            #Ask user to test another animal
            testAnotheranimal=self.validInput(input("\nTest a new animal?(Yes/No/Quit)"))
            #If the User Answers Yes
            if (testAnotheranimal==1): 
                beginProgram=1 #Jump back to start of loop and create new animal

            #If the User Answers No, print and export results, then exit the program
            if ((testAnotheranimal==0) or (testAnotheranimal==-1)): 
                results.printAll()
                workbookName=self.exportingResults(results)
                beginProgram=0 #Exit loop
                        
        #Exit program when testing is stopped
        if (beginProgram==0 ) or (beginProgram==-1):     
            print("*****Ending Testing*****")
        return workbookName

