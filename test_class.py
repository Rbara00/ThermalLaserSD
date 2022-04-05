###########################################
## 
## Table Class for thermal laser
## SD Team 10: Robert Bara, Zari Grandy, Ezra Galapo, Tyiana Smith
## 
## Author: Robert Bara
## Date:    4/5/2022
## Version: 1.1
##
## Description:
##  This python file contains a class to construct a table which will hold store
##  an animal into a list avaliable at the given index which is based upon a group,
##  graphically the table can be drawn as:
##      [column index][list of cats]
##       group0     : cat1,cat2,cat3
##       group1     : cat1,cat2,cat3
##        ...
##       groupN     : N cats
## Usage:
##  This program is invoked by test_class and utilizes inputs from the animal_class
##
########################################
import tables as resultTb
import animal_class as animal
from datetime import date   
from datetime import datetime
import os
from os import stat
from re import S
import re
from sre_constants import MAXGROUPS             # Import date class from datetime module
from statistics import pstdev    # Import statistics to utilize mean, standard deviation, and variance calculations
from statistics import pvariance
from statistics import mean
from enum import Enum
#import RPi.GPIO as GPIO        # Import General-Purpose In/Out for RPI4 to control laser and photodiode
from time import time
from numpy import empty, save           # Import time to calculate withdrawal time
import openpyxl                 # Import for exporting the animal's results to spreadsheets
#####################################################################
# Create a class to run a test and export all the animals information 
#####################################################################
class testAnimals:
    #No need for a constructor
    def __init__(self):
        self.numberOfGroups=0   #Keep track of the amount of groups avaliable
        self.numberOfAnimals=0  #Keep track of how many animals are tested upon
        return

    #####################################################
    # Functions for exporting results to spreadsheet
    #####################################################
    def exportingResults(self,results):
        resultTb=results
        workbookName=input("Save Output File As .xlsx?")
        #Check if the excel file exists, if not create it
        try: 
            #load the excel file with the corresponding workbook name
            wb1=openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        except FileNotFoundError:
            #Create the excel file with the corresponding workbook name
            wb1=self.exportSetup(workbookName)
            #Create enough sheets for each group
            for i in range(self.numberOfGroups+1):
                ws=wb1.create_sheet()
                ws.title="Group"+str(i)
                ws.append(["Group","Num of Group Trials","Group Avg","Animal Name","Trial Num","Withdrawal Times","Avg Time",
                            "Stdev","Trial Variance","Test Date", "Start Test Time","End Test Time"])
        
        #For 0 to max number of groups, load and write data from the results table into excel
        for i in range(self.numberOfGroups+1):
            curr_group=resultTb.arr[i]              #For readability, get the current group
            #print(wb1.sheetnames)
            groupTimes=[]                           #Keeps track of all withdrawawl times in the group
            groupTrialNum=1                         #Counter for how many trials were within a group
            #check if in the correct group sheet
            if ("Group"+str(i)) in wb1.sheetnames:
                ws=wb1["Group"+str(i)]              #Get the correct group's sheet
                #Iterate across the list at the group index, where each element is a stored animal
                for j in range(len(curr_group)):
                    curr_animal=curr_group[j]       #For readability, get the current animal
                    trialNum=1                      #For printing the trial number corresponding to the withdrawal time
                    analyzeTime=[]                  #For performing analysis for all data up to a given point
                    for w_time in curr_animal.time:
                        analyzeTime.append(w_time)  #store time local to cat for analysis
                        groupTimes.append(w_time)   #store time local to group for analysis
                        #Print all the group and cat data to excel
                        ws.append([curr_animal.getGroup(),groupTrialNum,mean(groupTimes),curr_animal.getName(), trialNum, w_time, 
                                mean(analyzeTime), pstdev(analyzeTime), pvariance(analyzeTime)
                                ,curr_animal.date, curr_animal.startTime,curr_animal.endTestTime])
                        trialNum+=1                 
                        groupTrialNum+=1
            #Otherwise create as many sheets as needed for the number of groups avaliable
            else:
                ws=wb1.create_sheet()
                #Label the sheet based on the group number
                ws.title="Group"+str(i)
                #Create column labels
                ws.append(["Group","Num of Group Trials","Group Avg","Animal Name","Trial Num","Withdrawal Times","Avg Time",
                        "Stdev","Trial Variance","Test Date", "Start Test Time","End Test Time"])
        #Save the excel file
        wb1.save(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        # Close the workbook and return
        wb1.close()
        return

    #This function creates a spreadsheet based on the workbook name and labels the proper columns for the animal's data
    def exportSetup(self,workbookName):
        #Create a work book
        wb1=openpyxl.Workbook()
        #Create a worksheet within the workbook
        ws1=wb1.active
        ws1.title="Results"
        #initialize all the column names
        ws1.append(["Group","Num of Group Trials","Group Avg","Animal Name","Trial Num","Withdrawal Times","Avg Time",
                     "Stdev","Trial Variance","Test Date", "Start Test Time","End Test Time"])
        #prompt for an output file name
        wb1.save(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbookName))
        #close the workbook
        wb1.close
        return wb1  #return the workbook so we can read/write to it with the other export functions
    
    ###############################################################################             
    #  Functions for starting the test and prompting for an animals information                          
    ###############################################################################    
    #Prompts the user for name and number to construct animal
    def promptAnimalInfo():
        #Prompt for name of Animal
        name=input("Enter subject's Name: ")
        #Prompt for a valid integer group number
        while True:
            try:
                group = int(input("Enter subject's Group number, minimum of 1: "))
                break
            except ValueError:
                print("Please enter Group number as an integer")  
            continue
        #Return Current Animal's name and number
        return name,group
    
    #Checks yes, no, or when to quit program
    def validInput(self,user_input):
        while True:
            if(user_input.lower()=="y") or user_input.lower()=="yes" or user_input=="1":
                output=1
                break
            if(user_input.lower()=="n") or user_input.lower()=="no" or user_input=="0":
                output=0
                break
            if(user_input.lower()=="q") or user_input.lower()=="quit" or user_input=="-1":
                output=-1
                break
            #prompt for a new input if invalid
            user_input=input("Please Enter a Valid Input (Yes/No/Quit)")
        return output

    #This method is what starts the entire system to even run a test
    def startExperiment(self):
        #Create a table to hold any results
        results=resultTb.resultTb()

       #Prompt user to start program
        print("\n*******Welcome to the Plantar Thermal Laser test.*******")
        beginProgram=self.validInput(input("Would you like to run a trial? (Yes/No/Quit): "))

        #Determine valid response, if so create a animal as an object
        while (beginProgram==1):
            #Create an animal data structure
            name,group=testAnimals.promptAnimalInfo()
            curr_animal=animal.animal(name,group)
            curr_animal.startTime=datetime.now().strftime("%H:%M:%S")  #get the timestamp when testing is beginning
            trialNum=1                                                  #Keeps track of trial of current animal

            #keep track of the maximum amount of groups
            if group>self.numberOfGroups:
                self.numberOfGroups=group 

            #Begin Testing a animal
            runTrial=1      #Flag to determine which animal is being tested on
            #Run as many trials as wanted for this animal
            while (runTrial==1):
                #Perform a Trial, get withdrawal time
                print("Trial ",trialNum,":")
                curr_animal.trial()

                #Prompt to perform another trial
                nextTrial=self.validInput(input("Run another trial for this animal?(Yes/No/Quit) "))
                if (nextTrial==1): #Keep testing current animal
                    trialNum+=1
                    runTrial=1 
                if (nextTrial==0): #Stop testing current animal
                    trialNum=0
                    runTrial=0
                if(nextTrial==-1): #Quit the program
                    beginProgram==-1
                    return

            #Print animal's information after testing 
            curr_animal.endTestTime=datetime.now().strftime("%H:%M:%S")  #get the timestamp when testing is over for cat
            curr_animal.printTime()

            #Insert the animal into the table based upon their group number
            results.insert(curr_animal)

            #Ask user to test another animal
            testAnotheranimal=self.validInput(input("\nTest a new animal?(Yes/No/Quit)"))
            #If the User Answers Yes
            if (testAnotheranimal==1): 
                beginProgram=1 #Jump back to start of loop and create new animal

            #If the User Answers No, print and export results, then exit the program
            if ((testAnotheranimal==0) or (testAnotheranimal==-1)): 
                results.printAll()
                self.exportingResults(results)
                beginProgram=0 #Exit loop
                        
        #Exit program when testing is stopped
        if (beginProgram==0 ) or (beginProgram==-1):     
            print("*****Ending Testing*****")
        return

#########################################
#           Driver Function             #
#########################################
def main():
    #create a test and start the experiment
    test=testAnimals()
    test.startExperiment()
    #exit program
    return 0

##Main definition##
if __name__ == "__main__":
    main()    